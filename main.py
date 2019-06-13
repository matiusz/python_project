#sala tabelka zajętości






import random
import math
import unicodedata
from openpyxl import *
import collections
import datetime
import time

class Room:
    @staticmethod
    #Make a room object with all fields empty - used to avoid crashes when handling the object
    def DeadRoom():
        room = Room("", "")
        room.classes = []
        return room
    #print the table of classroom reservations per day and time
    def printRoomTable(self):
        blocks = []
        weekdays = ["Pn", "Wt", "Sr", "Cz", "Pt"]
        blocks.append((datetime.time(8,0,0), datetime.time(9,30,0), "8.00"))
        blocks.append((datetime.time(9,35,0), datetime.time(11,5,0), "9.35"))
        blocks.append((datetime.time(11,15,0), datetime.time(12,45,0), "11.15"))
        blocks.append((datetime.time(12,50,0), datetime.time(14,20,0), "12.50"))
        blocks.append((datetime.time(14,40,0), datetime.time(16,10,0), "14.40"))
        blocks.append((datetime.time(16,15,0), datetime.time(17,45,0), "16.15"))
        blocks.append((datetime.time(17,50,0), datetime.time(19,20,0), "17.50"))
        blocks.append((datetime.time(19,30,0),datetime.time(21,0,0), "19.30"))
        print("\t", end = "\t|")
        for day in weekdays:
            print(day, end="\t|")
        print("")
        #print("-----------------------------")
        for i in range(8):
            print(blocks[i][2], end="\t| ")
            for day in weekdays:
                print(self.isTaken(day, blocks[i][0], blocks[i][1]), end="\t| ")
            print("")
            #print("-----------------------------")
    #given day and time range, returns X if the classroom is reserved in that time
    def isTaken(self, day, timeStart, timeEnd):
        for clas in self.classes:
            if (clas.day == day):
                if (clas.hourStart==""):
                    continue
                if ((timeStart <= clas.hourStart <= timeEnd)):
                    return clas.week or "X"
        return ""
    @staticmethod
    #return the room given string of "[building] [number]"
    def findByString(roomString, RoomsList):
        roomStringList = roomString.split()
        if (len(roomStringList)<2):
            return Room.DeadRoom()
        for room in RoomsList:
            if(room.building==roomStringList[0]):
                if(room.number==roomStringList[1]):
                    return room
        return Room.DeadRoom()
    def __init__(self, Building, Number):
        if (Building != None):
            self.building = Building
            self.classes = []
        else:
            self.building = ""
        if (Number != None):
            self.number = Number
        else:
            self.number = ""
    def __str__(self):
        return(self.building+" "+self.number)


class Person:
    @staticmethod
    #returns list of Persons with names similiar to name given
    def findSimiliarNames(string, PersonList, n=5):
        
        suggestedExactNames = [] #exact matches
        suggestedSimilarNames = [] #not exact matches
        split_string = string.split(" ")
        if len(split_string) == 1:
            for persons in PersonList:
                if string == persons.lastname[:len(string)]:
                    suggestedExactNames.append(persons)
                    continue
                dist = distance(string, persons.lastname)
                if(dist <= 2): # Similar lastnames but not exact
                    if(dist == 0): # Exact lastnames, check for similar firstnames
                        suggestedExactNames.append(persons)
                    else:
                        suggestedSimilarNames.append(persons)
        else:
            for persons in PersonList:
                if(distance(string, persons.lastname+" "+persons.firstname) == 0 or distance(string, persons.firstname+" "+persons.lastname) == 0): # Exact match
                    return [(1, persons)]
                if distance(split_string[0], persons.lastname) == 0:
                    if split_string[1] == persons.firstname[:len(split_string[1])]: #Exact lastname, exact beginning of firstname
                        suggestedExactNames.append(persons)
                        continue
                    if(distance(split_string[1], persons.firstname) <= 2): # Similar firstnames but not exact     
                        suggestedSimilarNames.append(persons)
                        continue
                if distance(split_string[0], persons.firstname) == 0:
                    if split_string[1] == persons.lastname[:len(split_string[1])]: #Exact lastname, exact beginning of firstname
                        suggestedExactNames.append(persons)
                        continue
                    if(distance(split_string[1], persons.lastname) <= 2): # Similar firstnames but not exact     
                        suggestedSimilarNames.append(persons)
                        continue
                if(distance(string, persons.lastname+" "+persons.firstname)<=n or distance(string, persons.firstname+" "+persons.lastname)<=n): # Similar match
                    suggestedSimilarNames.append(persons)
        if suggestedExactNames != []:
            return list(enumerate(suggestedExactNames, 1))
        if suggestedSimilarNames != []:
            return list(enumerate(suggestedSimilarNames, 1))

    @staticmethod
    #returns Person object with emmpty fields
    def DeadPerson():
        person = Person("", "")
        person.degree = ""
        person.classes = []
        return person
    @staticmethod
    #returns Person given it's exact first and last name
    def findByName(LastName, FirstName, PersonList):
        for person in PersonList:
            if(LastName==person.lastname):
                if (FirstName==person.firstname):
                    return person
        return Person.DeadPerson()
    def __init__(self, LastName, FirstName):
        self.firstname = FirstName
        self.lastname = LastName
        self.degree = ""
        self.classes = []
    def __str__(self):
        return(self.firstname+" "+self.lastname+", "+self.degree)


class Classes:
    def __init__(self, Subject):
        self.subject = Subject
    def __str__(self):
        return(self.subject+", "+str(self.person)+", "+str(self.classroom)+", "+self.day+" "+str(self.hourStart))
class Parser:
    @staticmethod
    #returns a list of Person objects from a given sheet
    def ParsePersons(persons):
        PersonList = []
        for row in persons.iter_rows(min_row=2):
            name = row[0].value.split()
            person = Person(name[0], name[1])
            PersonList.append(person)
            person.department = row[1].value
            person.job = row[2].value
            if (row[3].value!=None):
                person.degree = row[3].value
            person.additional = row[4].value
            person.pensum = row[5].value
            person.discount = row[6].value
            person.email = row[7].value
            person.room = row[8].value
            person.day = row[9].value
            person.time = row[10].value
        return PersonList
    #returns a list of classroom objects from a given sheet
    def ParseClassroom(classrooms):
        RoomsList = []
        for row in classrooms.iter_rows(min_row=2):
            if (row[0].value==row[1].value):
                continue
            classroom = Room(row[0].value, row[1].value)
            RoomsList.append(classroom)
            classroom.type = row[2].value
            classroom.capacity = row[3].value
            classroom.notes = row[4].value
        return RoomsList
    #returns a list of classes from a given sheet, adds them to matching persons and romm from the given lists
    def ParseSchedule(PersonList, RoomsList, sheet):
        ClassesList = []
        for row in sheet.iter_rows(min_row=2):
            if(row[0].value=="---" or row[0].value==None or row[3].value == None):
                continue
            classes = Classes(row[3].value)
            ClassesList.append(classes)
            classes.course = row[0].value
            if(row[1]!=None):
                classes.semester = row[1].value
            else:
                classes.semester = ""
            if(row[2].value != None):
                classes.location = row[2].value
            else:
                classes.location = ""
            if (row[4].value != None):
                classes.chooseable = row[4].value
            else:
                classes.chooseable = row[5].value
            if (row[5].value!=None):
                classes.type = row[5].value
            else:
                classes.type = ""
            if (row[6].value != None):
                classes.group = row[6].value
            else:
                classes.group = ""
            if(row[7].value!=None):
                classes.timeframe = row[7].value
            else:
                classes.timeframe = ""
            if(row[9].value!=None):
                classes.faculty = row[9].value
            else:
                classes.faculty = ""
            nameString = row[10].value
            if (nameString!=None):
                name=nameString.split()
                classes.person = Person.findByName(name[0], name[1], PersonList)
                classes.person.classes.append(classes)
            else:
                classes.person = Person.DeadPerson()
            classroomString = row[11].value
            if (classroomString!=None):
                classes.classroom = Room.findByString(classroomString, RoomsList)
                classes.classroom.classes.append(classes)
            else:
                classes.classroom = Room.DeadRoom()
            if (row[12].value!=None):
                classes.week = row[12].value
            else:
                classes.week = ""
            if(row[12].value!=None):
                classes.week = row[12].value or ""
            else:
                classes.week
            classes.day = row[13].value or ""
            classes.hourStart = row[14].value or ""
            classes.hourEnd = row[15].value or ""
        return ClassesList
    @staticmethod
    #load a file, parse all fields
    def ParseOther(RoomsList, sheet):
        OtherList = []
        for row in sheet.iter_rows(min_row=2):
            classes = Classes(row[1].value)
            OtherList.append(classes)
            classroomString = row[2].value
            if (classroomString!=None):
                classes.classroom = Room.findByString(classroomString, RoomsList)
                classes.classroom.classes.append(classes)
            classes.week = ""
            classes.type = "Other"
            classes.person = Person.DeadPerson()
            classes.day = row[3].value or ""
            classes.hourStart = row[4].value or ""
            classes.hourEnd = row[5].value or ""
        return OtherList


    def ParseAll(filename = '2019-2020(6) (1).xlsx', option = "oba"):
        workbook = load_workbook(filename)
        personsSheet = workbook["osoby"]
        roomsSheet = workbook["sale"]
        PersonsList = Parser.ParsePersons(personsSheet)
        RoomsList = Parser.ParseClassroom(roomsSheet)
        OthersWinterList = []
        OthersSummerList = []
        if (option == "oba" or option == "zima"):
            OthersWinterList = Parser.ParseOther(RoomsList, workbook["zima_inne"])
        if (option == "oba" or option == "lato"):
            OthersSummerList = Parser.ParseOther(RoomsList, workbook["lato_inne"])
        SummerNList = []
        SummerSList = []
        WinterNList = []
        WinterSList = []
        if (option == "oba" or option == "lato"):
            SummerSList = Parser.ParseSchedule(PersonsList, RoomsList, workbook["lato_s"])
            SummerNList = Parser.ParseSchedule(PersonsList, RoomsList, workbook["lato_n"])
        if (option == "oba" or option == "zima"):
            WinterNList = Parser.ParseSchedule(PersonsList, RoomsList, workbook["zima_n"])
            WinterSList = Parser.ParseSchedule(PersonsList, RoomsList, workbook["zima_s"])
        AllClasses = WinterSList+SummerSList+WinterNList+SummerNList+OthersSummerList+OthersWinterList
        return (PersonsList, RoomsList, AllClasses) # check if tuple order matches

        
def printPersonData(person):
    print("\n" + str(person))
    print("Kontakt: " + (person.email or ""))
    print("Konsultacje: " + (person.room or "")+ " " + (person.day or "")+ " " + str(person.time or ""))
    print("Prowadzone zajęcia: ")
    classes_monday = []
    classes_tuesday = []
    classes_wednesday = []
    classes_thursday = []
    classes_friday = []
    classes_other = []
    '''for classes in person.classes:
        print(classes.subject + " " + str(classes.classroom) + " " + classes.day + " " + str(classes.hourStart))'''
    for classes in person.classes:
        if classes.day == "Pn":
            classes_monday.append(classes)
        elif classes.day == "Wt":
            classes_tuesday.append(classes)
        elif classes.day == "Sr":
            classes_wednesday.append(classes)
        elif classes.day == "Cz":
            classes_thursday.append(classes)
        elif classes.day == "Pt":
            classes_friday.append(classes)
        else:
            classes_other.append(classes)
    print("\tPoniedziałek:")
    if classes_monday == []:
            print("Brak zajęć")
    else:
        for classes in classes_monday:
            print(classes.subject + " " + str(classes.classroom) + " " + str(classes.hourStart))
    print("\tWtorek:")
    if classes_tuesday == []:
            print("Brak zajęć")
    else:
        for classes in classes_tuesday:
            print(classes.subject + " " + str(classes.classroom) + " " + str(classes.hourStart))
    print("\tŚroda:")
    if classes_wednesday == []:
            print("Brak zajęć")
    else:
        for classes in classes_wednesday:
            print(classes.subject + " " + str(classes.classroom) + " " + str(classes.hourStart))
    print("\tCzwartek:")
    if classes_thursday == []:
            print("Brak zajęć")
    else:
        for classes in classes_thursday:
            print(classes.subject + " " + str(classes.classroom) + " " + str(classes.hourStart))
    print("\tPiątek:")
    if classes_friday == []:
            print("Brak zajęć")
    else:
        for classes in classes_friday:
            print(classes.subject + " " + str(classes.classroom) + " " + str(classes.hourStart))
    print("\tPozostałe:")
    if classes_other == []:
            print("Brak zajęć")
    else:
        for classes in classes_other:
            print(classes.subject + " " + str(classes.classroom) + " " + classes.day + " " + str(classes.hourStart))

# choose person from list of suggestions and call printPersonData
def getPerson(arg, PersonsList):
    suggestions =  Person.findSimiliarNames(arg,PersonsList)
    if suggestions == None:
        print("Nie znaleziono podanej osoby")
    elif len(suggestions) == 1:
        printPersonData(suggestions[0][1])
    else:
        for index, suggestion in suggestions:
            print(str(index) + ": " + suggestion.lastname+" "+suggestion.firstname)
        print("Wprowadź numer odpowiadający osobie")
        i = int(input())
        if i <= len(suggestions) and i > 0:
            printPersonData(suggestions[i-1][1])
#calculete Levenstein's distance
def distance(a,b):
    n, m = len(a), len(b)
    if n > m:
        a,b = b,a
        n,m = m,n
    current = range(n+1)
    for i in range(1,m+1):
        previous, current = current, [i]+[0]*m
        for j in range(1,n+1):
            add, delete = previous[j]+1, current[j-1]+1
            change = previous[j-1]
            if a[j-1] != b[i-1]:
                change = change + 1
            current[j] = min(add, delete, change)
    return current[n]


def getRoom(arg, RoomsList):
    suggestions = []
    for room in RoomsList:
        if (room.building=="D17"):
            if (room.number==arg):
                room.printRoomTable()
                return room
            else:
                if (distance(arg, room.number)<=1):
                    suggestions.append(room)
    suggs = []
    for suggestion in suggestions:
        if (len(suggestion.number)==5):
            suggs.append(suggestion)
    if (suggs!=[]):
        suggestions = suggs
    suggestionsEnumerated = list(enumerate(suggestions, 1))
    if (suggestions == None):
        print("Nie znaleziono podanej sali")
    else:
        for index, suggestion in suggestionsEnumerated:
            print(str(index) + ": " + suggestion.number)
        print("Wprowadź numer odpowiadający sali")
        i = int(input())
        if i <= len(suggestions) and i > 0:
            suggestionsEnumerated[i-1][1].printRoomTable()



def Main():
    print("Opcje: \noba - załaduj wszytkie zajęcia\nlato - załaduj semestr letni\nzima - załaduj semestr zimowy")
    print(">>", end='')
    input1 = input()
    print("Parsowanie danych...")
    data = Parser.ParseAll(option = input1)
    print("Parsowanie ukończone")
    print(">>", end='')
    input1 = input()
    split_input =  input1.split(" ", 1)
    function = split_input[0]
    if (len(split_input)>1):
        arg = split_input[1]
    else:
        print("osoba [Nazwisko] ")
        print("sala [Numer sali]")
    while function.lower() != "stop":
        if function.lower() == "osoba":
            getPerson(arg, data[0])
        if function.lower() == "sala":
            getRoom(arg, data[1])
        print(">>", end='')
        input1 = input()
        split_input =  input1.split(" ", 1)
        function = split_input[0]
        arg = split_input[1]

if __name__=="__main__":
    Main()